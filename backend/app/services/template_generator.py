from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# (header label, mandatory, example value). Header labels intentionally match
# header_mapper.CANONICAL_HEADERS' canonical variants closely, so a template
# filled in and re-uploaded via POST /{event_id}/upload auto-maps with zero
# unmapped_headers.
#
# "mandatory" is a template-level UX nudge, not a strict mirror of
# validation.validate_rows's reject conditions. Name/Email/Company are
# reject-if-missing there, and are marked mandatory here for that reason.
# Looking For/Offerings are also marked mandatory here to push uploaders
# toward filling them in - a row that skips both still only gets *flagged*
# for admin review by validate_rows, not rejected (CLAUDE.md: "Sparse rows
# are flagged, not dropped"). If validate_rows' reject conditions ever
# change, double check whether this list should follow.
TEMPLATE_COLUMNS: list[tuple[str, bool, str]] = [
    ("Name", True, "Jane Doe"),
    ("Email", True, "jane.doe@example.com"),
    ("Company", True, "Acme Corp"),
    ("Designation", False, "Marketing Director"),
    ("Sector", False, "Technology"),
    ("Company Size", False, "50-200"),
    ("Membership Tier", False, "Premium Member"),
    ("Looking For", True, "Investors for a Series A round"),
    ("Offerings", True, "SaaS analytics platform for retail"),
    ("Ideal Connection", False, "Corporate innovation leads"),
    ("Biggest Opportunity", False, "Expanding into the DACH market"),
    ("Website", False, "https://acme.example.com"),
    ("LinkedIn URL", False, "https://linkedin.com/in/janedoe"),
    ("Phone", False, "+31 6 1234 5678"),
]

# Light amber - flags mandatory columns to a sighted user even without
# relying on bold alone (e.g. at a glance, or a low-contrast display).
MANDATORY_HEADER_FILL = "FFF2CC"
EXAMPLE_ROW_FONT_COLOR = "999999"

# Dropdown options for the Membership Tier column. Each of these strings is
# individually confirmed to normalize cleanly via
# tier_normalizer.KNOWN_TIER_VARIANTS (case/punctuation-insensitive) to its
# obvious MembershipTier - "Non-Member" -> "non member" -> non_member, etc.
# So a value picked from this dropdown always lands in valid/eligible, never
# in the "unrecognized -> business_member, flagged for review" fallback path.
MEMBERSHIP_TIER_OPTIONS = ["Sponsor", "Premium Member", "Business Member", "Normal Member", "Non-Member"]

# How many data rows (below the header) get the dropdown applied - generous
# headroom for a real participant list, not tied to any actual row count.
DATA_VALIDATION_ROW_COUNT = 1000


def generate_participant_template() -> bytes:
    """Build a blank participant-upload .xlsx.

    Row 1: column headers, mandatory ones bold + highlighted. Row 2: a
    greyed-out italic example row showing the expected shape/format per
    column - not real data, just a fill-in guide; the ingestion pipeline
    would reject it as a real row only if actually uploaded verbatim (real
    email format, so it would in fact pass validation - left in deliberately
    as a working example rather than obvious placeholder junk).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    example_font = Font(italic=True, color=EXAMPLE_ROW_FONT_COLOR)
    mandatory_fill = PatternFill("solid", fgColor=MANDATORY_HEADER_FILL)

    for col_idx, (label, mandatory, _example) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = bold_font if mandatory else normal_font
        if mandatory:
            cell.fill = mandatory_fill

    for col_idx, (_label, _mandatory, example) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=example)
        cell.font = example_font

    for col_idx, (label, *_rest) in enumerate(TEMPLATE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(label) + 4)

    tier_col_idx = next(i for i, (label, *_) in enumerate(TEMPLATE_COLUMNS, start=1) if label == "Membership Tier")
    tier_col_letter = get_column_letter(tier_col_idx)
    dv = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(MEMBERSHIP_TIER_OPTIONS)),
        allow_blank=True,
        # openpyxl/Excel XML quirk: showDropDown is inverted from what its name
        # suggests - True actually HIDES the in-cell dropdown arrow, False (or
        # omitted) is what shows it. Deliberately explicit here so it doesn't
        # get "fixed" to True by someone reading the name literally later.
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Membership Tier",
        error="Please pick one of: " + ", ".join(MEMBERSHIP_TIER_OPTIONS),
        promptTitle="Membership Tier",
        prompt="Choose one from the dropdown, or leave blank for Normal Member (the default).",
    )
    ws.add_data_validation(dv)
    dv.add(f"{tier_col_letter}2:{tier_col_letter}{DATA_VALIDATION_ROW_COUNT}")

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
