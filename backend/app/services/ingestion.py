import csv
import io
import logging
from dataclasses import dataclass, field
from typing import Iterable, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from app.services.header_mapper import map_headers, map_row
from app.services.tier_normalizer import normalize_tier
from app.services.validation import validate_rows

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = (".xlsx", ".csv")

# Participant fields populated directly from the mapped/validated row.
# membership_tier is set separately after tier normalization.
_PARTICIPANT_FIELD_KEYS = (
    "name", "email", "company", "designation", "website", "linkedin_url",
    "phone", "sector", "company_size", "looking_for", "offerings",
    "ideal_connection", "biggest_opportunity",
)


def parse_upload(file_bytes: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    """Parse an uploaded .xlsx or .csv file into raw row dicts.

    Returns (valid_rows, skipped_rows). Never raises on a malformed row —
    row-level failures are logged and collected in skipped_rows instead.
    """
    lower_name = filename.lower()

    if lower_name.endswith(".xlsx"):
        return _parse_xlsx(file_bytes)
    if lower_name.endswith(".csv"):
        return _parse_csv(file_bytes)

    raise ValueError(
        f"Unsupported file type for '{filename}'. Supported formats: {', '.join(SUPPORTED_EXTENSIONS)}"
    )


def _clean_headers(raw_headers: Iterable) -> list[str]:
    return [str(h).strip() if h is not None else "" for h in raw_headers]


def _row_is_empty(values) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _build_row_dict(headers: list[str], values) -> dict:
    row_dict = {}
    for header, value in zip(headers, values):
        if not header:
            continue
        row_dict[header] = value.strip() if isinstance(value, str) else value
    return row_dict


def _parse_xlsx(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    valid_rows: list[dict] = []
    skipped_rows: list[dict] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Failed to open xlsx file: {e}")
        return [], [{"row_number": None, "reason": f"Could not open file: {e}"}]

    if ws is None:
        logger.error("Workbook has no active worksheet")
        return [], [{"row_number": None, "reason": "Workbook has no active worksheet"}]

    # Propagate merged-cell values across their full range so every cell
    # in a merged block carries data instead of being blank.
    try:
        for merged_range in list(ws.merged_cells.ranges):
            top_left_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    # unmerge_cells() above already converted every cell in this
                    # range from MergedCell to a plain, writable Cell.
                    cast(Cell, ws.cell(row=row, column=col)).value = top_left_value
    except Exception as e:
        logger.warning(f"Could not fully unmerge cells: {e}")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return [], [{"row_number": None, "reason": "File is empty"}]

    headers = _clean_headers(raw_headers)

    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or _row_is_empty(row):
            continue

        try:
            row_dict = _build_row_dict(headers, row)
            row_dict["__row_number"] = row_number
            valid_rows.append(row_dict)
        except Exception as e:
            logger.warning(f"Skipping row {row_number}: {e}")
            skipped_rows.append({"row_number": row_number, "reason": str(e)})

    return valid_rows, skipped_rows


@dataclass
class ParticipantRecord:
    row_number: int | None
    fields: dict
    participant_status: str
    raw_source_data: dict
    reasons: list[str] = field(default_factory=list)


@dataclass
class IngestionPipelineResult:
    total_rows: int
    parse_skipped: int
    unmapped_headers: list[str]
    valid_count: int
    flagged_count: int
    rejected_count: int
    rejected_details: list[dict]
    participants: list[ParticipantRecord]


def run_ingestion_pipeline(file_bytes: bytes, filename: str) -> IngestionPipelineResult:
    """Full ingestion chain: parse -> map headers -> validate -> normalize tier.

    Returns everything the API layer needs to bulk-insert Participant rows,
    without touching the database itself.
    """
    raw_rows, skipped_rows = parse_upload(file_bytes, filename)

    if not raw_rows:
        return IngestionPipelineResult(
            total_rows=0,
            parse_skipped=len(skipped_rows),
            unmapped_headers=[],
            valid_count=0,
            flagged_count=0,
            rejected_count=0,
            rejected_details=[],
            participants=[],
        )

    headers = [h for h in raw_rows[0].keys() if not h.startswith("__")]
    header_map, unmapped_headers = map_headers(headers)
    mapped_rows = [map_row(r, header_map) for r in raw_rows]
    result = validate_rows(mapped_rows)

    raw_by_row_number = {
        r["__row_number"]: {k: v for k, v in r.items() if not k.startswith("__")}
        for r in raw_rows
    }

    participants: list[ParticipantRecord] = []

    for bucket, needs_review in ((result.valid, False), (result.flagged, True)):
        for item in bucket:
            tier_result = normalize_tier(item.data.get("membership_tier"))
            participant_status = "review" if (needs_review or tier_result.flagged) else "eligible"

            fields = {key: item.data.get(key) for key in _PARTICIPANT_FIELD_KEYS}
            fields["membership_tier"] = tier_result.tier

            reasons = list(item.reasons)
            if tier_result.flagged and tier_result.reason:
                reasons.append(tier_result.reason)

            participants.append(
                ParticipantRecord(
                    row_number=item.row_number,
                    fields=fields,
                    participant_status=participant_status,
                    raw_source_data=raw_by_row_number.get(item.row_number, {}),
                    reasons=reasons,
                )
            )

    return IngestionPipelineResult(
        total_rows=len(raw_rows),
        parse_skipped=len(skipped_rows),
        unmapped_headers=unmapped_headers,
        valid_count=len(result.valid),
        flagged_count=len(result.flagged),
        rejected_count=len(result.rejected),
        rejected_details=[{"row_number": r.row_number, "reasons": r.reasons} for r in result.rejected],
        participants=participants,
    )


def _decode_csv_bytes(file_bytes: bytes) -> str | None:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def _parse_csv(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    valid_rows: list[dict] = []
    skipped_rows: list[dict] = []

    text = _decode_csv_bytes(file_bytes)
    if text is None:
        return [], [{"row_number": None, "reason": "Could not decode file with any supported encoding"}]

    try:
        reader = csv.reader(io.StringIO(text))
        raw_headers = next(reader)
    except StopIteration:
        return [], [{"row_number": None, "reason": "File is empty"}]
    except Exception as e:
        logger.error(f"Failed to parse csv file: {e}")
        return [], [{"row_number": None, "reason": f"Could not parse CSV: {e}"}]

    headers = _clean_headers(raw_headers)

    for row_number, row in enumerate(reader, start=2):
        if not row or _row_is_empty(row):
            continue

        try:
            row_dict = _build_row_dict(headers, row)
            row_dict["__row_number"] = row_number
            valid_rows.append(row_dict)
        except Exception as e:
            logger.warning(f"Skipping row {row_number}: {e}")
            skipped_rows.append({"row_number": row_number, "reason": str(e)})

    return valid_rows, skipped_rows
